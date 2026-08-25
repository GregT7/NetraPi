"""Rebuild per-clip ap_050 labels + kNN features from the local motion_area_2 cache.

Does not need the missing analysis .py tree. Uses:
  - tagged clip ids from vids/unsafe_events/clip_docs_compr.xlsx
  - cached *.areas.json / *.motion.json (default: this folder's data/ copy)
  - edge winner_pf02 approach + runtime feature extractors
  - frozen knn_stage1/2.joblib for optional predictions

Usage (from repo root):

    python src/tests/integration/at_3_4/export_ap_050_clip_features.py
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[3]
EDGE_DIR = REPO_ROOT / "src" / "main" / "edge"
MAIN_DIR = REPO_ROOT / "src" / "main"
CONFIG_DIR = SCRIPT_DIR / "config"
DEFAULT_DATA_DIR = SCRIPT_DIR / "data"
DEFAULT_TAGS = REPO_ROOT / "vids" / "unsafe_events" / "clip_docs_compr.xlsx"
DEFAULT_OUT = CONFIG_DIR / "ap_050_clip_features.json"
DEFAULT_POINTS_OUT = (
    REPO_ROOT / "src" / "main" / "frontend" / "src" / "components" / "ap050ClipPoints.json"
)
DEFAULT_PCA_OUT = (
    REPO_ROOT / "src" / "main" / "frontend" / "src" / "components" / "ap050Stage1Pca.json"
)
DEFAULT_STAGE1_FEATURES_OUT = (
    REPO_ROOT / "src" / "main" / "frontend" / "src" / "components" / "ap050Stage1Features.json"
)

CLIP_ID_RE = re.compile(r"clip_(\d+)_", re.IGNORECASE)

TAG_TO_CATEGORY = {
    "run_through_stop": "run-through",
    "complete_stop": "complete-stop",
    "rolling_stop": "rolling-stop",
    "irrelevant_stop_sign": "unrelated",
    "stop_queue": "unrelated",
    "unrelated": "unrelated",
}

TAG_PRIORITY = (
    "run_through_stop",
    "complete_stop",
    "rolling_stop",
    "irrelevant_stop_sign",
    "stop_queue",
    "unrelated",
)

DISPLAY_LABEL = {
    "complete-stop": "Complete stop",
    "rolling-stop": "Rolling stop",
    "run-through": "Run-through",
    "unrelated": "Unrelated",
}

STAGE1_CHART_LABEL = {
    "complete-stop": "Complete stop",
    "rolling-stop": "Rolling / run-through",
    "run-through": "Rolling / run-through",
}
STAGE2_CHART_LABELS = frozenset({"rolling-stop", "run-through"})

STAGE1_NAMES = (
    "post_drop_mean_motion",
    "post_drop_min_motion",
    "post_drop_p95_motion",
    "post_drop_stop_fraction",
)
STAGE2_NAMES = (
    "post_drop_min_motion",
    "approach_area_sum_pct",
)


def _bootstrap_edge() -> None:
    for path in (str(MAIN_DIR), str(EDGE_DIR), str(SCRIPT_DIR)):
        if path not in sys.path:
            sys.path.insert(0, path)


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def clip_id_from_name(name: str) -> int | None:
    match = CLIP_ID_RE.search(name)
    return int(match.group(1)) if match else None


def load_tags(xlsx_path: Path) -> dict[int, tuple[str, str]]:
    """Return clip_id -> (raw_tag, category label)."""
    import openpyxl

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "tags" not in workbook.sheetnames:
        raise SystemExit(f"No 'tags' sheet in {xlsx_path}")

    by_id: dict[int, list[str]] = {}
    for row in workbook["tags"].iter_rows(min_row=2, values_only=True):
        raw_id, raw_tag = row[0], row[1]
        if raw_id is None or raw_tag is None:
            continue
        clip_id = int(raw_id)
        tag = str(raw_tag).strip()
        by_id.setdefault(clip_id, []).append(tag)

    labels: dict[int, tuple[str, str]] = {}
    for clip_id, tags in by_id.items():
        chosen = next((tag for tag in TAG_PRIORITY if tag in tags), tags[0])
        category = TAG_TO_CATEGORY.get(chosen, chosen)
        labels[clip_id] = (chosen, category)
    return labels


def named_features(names: tuple[str, ...], values: list[float] | None) -> dict[str, float] | None:
    if values is None:
        return None
    return {name: float(value) for name, value in zip(names, values, strict=True)}


def first_prefix_detect(
    areas: list[float],
    fps: float,
    approach_config,
) -> tuple[int, float] | None:
    from netrapi.events.approach import diagnose_approach_drop

    if not areas or fps <= 0:
        return None
    min_frames = max(2, int(approach_config.min_approach_s * fps))
    for frame in range(min_frames, len(areas)):
        diagnosis = diagnose_approach_drop(
            areas[: frame + 1],
            fps,
            config=approach_config,
        )
        if diagnosis is not None and diagnosis.event is not None:
            return frame, frame / fps
    return None


def load_series(areas_path: Path) -> tuple[list[float], list[float], float, str]:
    areas_doc = _load_json(areas_path)
    motion_path = areas_path.with_name(areas_path.name.replace(".areas.json", ".motion.json"))
    if not motion_path.is_file():
        raise FileNotFoundError(motion_path)
    motion_doc = _load_json(motion_path)

    areas = [float(value) for value in areas_doc["areas"]]
    scores = [float(value) for value in motion_doc["scores"]]
    fps = float(
        motion_doc.get("source_video", {}).get("fps")
        or areas_doc.get("source_video", {}).get("fps")
        or 0.0
    )
    stem = str(areas_doc.get("maps_to_clip") or areas_path.name).removesuffix(".mp4")
    return areas, scores, fps, stem


def predict_knn(
    stage1_features: list[float],
    stage2_features: list[float],
    classifier,
) -> dict[str, str]:
    from at_3_4_pipeline import e2e_predicted

    stage1 = str(classifier._stage1.predict([stage1_features])[0])
    stage2 = ""
    if stage1 == "rolling-or-run-through":
        stage2 = str(classifier._stage2.predict([stage2_features])[0])
    return {
        "predicted_stage1": stage1,
        "predicted_stage2": stage2,
        "predicted_e2e": e2e_predicted(stage1, stage2),
    }


def stage2_scatter_points(clips: list[dict[str, Any]]) -> list[dict[str, Any]]:
    points = []
    for row in clips:
        stage2 = row.get("stage2_features")
        if not stage2 or row.get("label") not in STAGE2_CHART_LABELS:
            continue
        points.append(
            {
                "clipId": row["clip_id"],
                "x": stage2["post_drop_min_motion"],
                "y": stage2["approach_area_sum_pct"],
                "label": row["label_display"],
            }
        )
    return points


def stage1_feature_rows(clips: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in clips:
        features = row.get("stage1_features")
        if not features or row.get("label") not in STAGE1_CHART_LABEL:
            continue
        rows.append(
            {
                "clipId": row["clip_id"],
                "label": STAGE1_CHART_LABEL[row["label"]],
                **{name: float(features[name]) for name in STAGE1_NAMES},
            }
        )
    return rows


def stage1_pca_payload(clips: list[dict[str, Any]]) -> dict[str, Any]:
    import numpy as np
    from sklearn.decomposition import PCA
    from sklearn.preprocessing import StandardScaler

    rows = [
        row
        for row in clips
        if row.get("stage1_features") and row.get("label") in STAGE1_CHART_LABEL
    ]
    if len(rows) < 2:
        raise SystemExit("Need at least 2 clips with stage-1 features for PCA.")

    matrix = np.asarray(
        [[float(row["stage1_features"][name]) for name in STAGE1_NAMES] for row in rows],
        dtype=float,
    )
    scaled = StandardScaler().fit_transform(matrix)
    pca = PCA(n_components=2, random_state=0)
    coords = pca.fit_transform(scaled)
    loadings = pca.components_
    return {
        "featureNames": list(STAGE1_NAMES),
        "explainedVariance": [float(value) for value in pca.explained_variance_ratio_],
        "loadings": {
            "pc1": {name: float(loadings[0, index]) for index, name in enumerate(STAGE1_NAMES)},
            "pc2": {name: float(loadings[1, index]) for index, name in enumerate(STAGE1_NAMES)},
        },
        "points": [
            {
                "clipId": row["clip_id"],
                "x": float(coords[index, 0]),
                "y": float(coords[index, 1]),
                "label": STAGE1_CHART_LABEL[row["label"]],
            }
            for index, row in enumerate(rows)
        ],
    }


def write_chart_json(
    clips: list[dict[str, Any]],
    *,
    points_out: Path,
    pca_out: Path,
    stage1_features_out: Path,
) -> tuple[int, int, int]:
    points = stage2_scatter_points(clips)
    pca = stage1_pca_payload(clips)
    stage1_rows = stage1_feature_rows(clips)
    points_out.parent.mkdir(parents=True, exist_ok=True)
    pca_out.parent.mkdir(parents=True, exist_ok=True)
    stage1_features_out.parent.mkdir(parents=True, exist_ok=True)
    points_out.write_text(json.dumps(points, indent=2) + "\n", encoding="utf-8")
    pca_out.write_text(json.dumps(pca, indent=2) + "\n", encoding="utf-8")
    stage1_features_out.write_text(json.dumps(stage1_rows, indent=2) + "\n", encoding="utf-8")
    return len(points), len(pca["points"]), len(stage1_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export ap_050 per-clip labels and kNN features.")
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--tags", type=Path, default=DEFAULT_TAGS)
    parser.add_argument("--config-dir", type=Path, default=CONFIG_DIR)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--points-out", type=Path, default=DEFAULT_POINTS_OUT)
    parser.add_argument("--pca-out", type=Path, default=DEFAULT_PCA_OUT)
    parser.add_argument("--stage1-features-out", type=Path, default=DEFAULT_STAGE1_FEATURES_OUT)
    parser.add_argument(
        "--from-json",
        type=Path,
        help="Skip clip extraction; rebuild scatter/PCA JSON from an existing features file.",
    )
    parser.add_argument("--skip-predict", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.from_json is not None:
        payload = _load_json(args.from_json)
        n_stage2, n_pca, n_stage1 = write_chart_json(
            payload["clips"],
            points_out=args.points_out,
            pca_out=args.pca_out,
            stage1_features_out=args.stage1_features_out,
        )
        print(f"Wrote {args.points_out} ({n_stage2} stage-2 points)")
        print(f"Wrote {args.pca_out} ({n_pca} stage-1 PCA points)")
        print(f"Wrote {args.stage1_features_out} ({n_stage1} stage-1 feature rows)")
        return 0

    _bootstrap_edge()

    from config.types import ApproachConfig
    from netrapi.events.classify.features import extract_stage1_features, extract_stage2_features

    if not args.data_dir.is_dir():
        print(f"Missing data dir: {args.data_dir}", file=sys.stderr)
        return 1
    if not args.tags.is_file():
        print(f"Missing tags workbook: {args.tags}", file=sys.stderr)
        return 1

    approach = ApproachConfig.from_json(_load_json(args.config_dir / "approach_config.json"))
    metrics = _load_json(args.config_dir / "metrics_config.json")
    motion = _load_json(args.config_dir / "motion_config.json")
    window_s = float(metrics["post_drop_window_s"])
    stopped_threshold = float(motion["stopped_motion_threshold"])
    labels_by_id = load_tags(args.tags)

    classifier = None
    if not args.skip_predict:
        try:
            from netrapi.events.classify import StopClassifier

            models = REPO_ROOT / "src" / "main" / "edge" / "models"
            classifier = StopClassifier.from_paths(
                models / "knn_stage1.joblib",
                models / "knn_stage2.joblib",
            )
        except Exception as exc:
            print(f"Skipping joblib predictions: {exc}", file=sys.stderr)

    clips: list[dict[str, Any]] = []
    areas_files = sorted(args.data_dir.glob("*.areas.json"))
    for areas_path in areas_files:
        clip_id = clip_id_from_name(areas_path.name)
        if clip_id is None:
            continue
        tag_row = labels_by_id.get(clip_id)
        areas, scores, fps, stem = load_series(areas_path)
        n = min(len(areas), len(scores))
        areas = areas[:n]
        scores = scores[:n]
        motion_samples = [(index / fps, scores[index]) for index in range(n)]

        row: dict[str, Any] = {
            "clip_id": clip_id,
            "file_stem": stem,
            "fps": fps,
            "frame_count": n,
            "tag": tag_row[0] if tag_row else None,
            "label": tag_row[1] if tag_row else None,
            "label_display": DISPLAY_LABEL.get(tag_row[1], tag_row[1]) if tag_row else None,
            "approach_fired": False,
            "detect_frame": None,
            "t0_s": None,
            "stage1_features": None,
            "stage2_features": None,
            "predicted_stage1": None,
            "predicted_stage2": None,
            "predicted_e2e": None,
        }

        detected = first_prefix_detect(areas, fps, approach)
        if detected is None:
            clips.append(row)
            continue

        detect_frame, t0_s = detected
        stage1 = extract_stage1_features(
            motion_samples,
            anchor_s=t0_s,
            window_s=window_s,
            stopped_threshold=stopped_threshold,
        )
        stage2 = None
        if stage1 is not None:
            stage2 = extract_stage2_features(
                stage1_features=stage1,
                areas_snapshot=areas,
                detect_frame=detect_frame,
                fps=fps,
                approach_config=approach,
            )

        row["approach_fired"] = True
        row["detect_frame"] = detect_frame
        row["t0_s"] = t0_s
        row["stage1_features"] = named_features(STAGE1_NAMES, stage1)
        row["stage2_features"] = named_features(STAGE2_NAMES, stage2)
        if classifier is not None and stage1 is not None and stage2 is not None:
            row.update(predict_knn(stage1, stage2, classifier))
        clips.append(row)

    matched_ids = {row["clip_id"] for row in clips}
    unmatched_tags = sorted(set(labels_by_id) - matched_ids)
    unlabeled_files = sorted(row["clip_id"] for row in clips if row["label"] is None)

    summary = {
        "areas_files": len(areas_files),
        "tagged_clips": len(labels_by_id),
        "exported_clips": len(clips),
        "approach_fired": sum(1 for row in clips if row["approach_fired"]),
        "stage1_ready": sum(1 for row in clips if row["stage1_features"]),
        "stage2_ready": sum(1 for row in clips if row["stage2_features"]),
        "by_label": dict(Counter(row["label"] for row in clips if row["label"])),
        "unmatched_tag_ids": unmatched_tags,
        "unlabeled_file_ids": unlabeled_files,
    }

    payload = {
        "experiment_id": "ap_050",
        "experiment_name": "window_5_motion_area_2_stopped_lo_k3_min_area",
        "approach_anchor": "per_frame_first_detect",
        "stage1_feature_names": list(STAGE1_NAMES),
        "stage2_feature_names": list(STAGE2_NAMES),
        "post_drop_window_s": window_s,
        "stopped_motion_threshold": stopped_threshold,
        "prediction_note": (
            "predicted_* columns use the frozen Pi joblibs on these same clips "
            "(in-sample), not the original leave-one-out 83.3% run."
        ),
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "source": {
            "data_dir": str(args.data_dir.relative_to(REPO_ROOT)).replace("\\", "/"),
            "tags_workbook": str(args.tags.relative_to(REPO_ROOT)).replace("\\", "/"),
            "approach_config": "winner_pf02",
        },
        "summary": summary,
        "clips": clips,
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    n_stage2, n_pca, n_stage1 = write_chart_json(
        clips,
        points_out=args.points_out,
        pca_out=args.pca_out,
        stage1_features_out=args.stage1_features_out,
    )

    print(f"Wrote {args.out}")
    print(f"Wrote {args.points_out} ({n_stage2} stage-2 scatter points)")
    print(f"Wrote {args.pca_out} ({n_pca} stage-1 PCA points)")
    print(f"Wrote {args.stage1_features_out} ({n_stage1} stage-1 feature rows)")
    print(
        f"  files={summary['areas_files']} tagged={summary['tagged_clips']} "
        f"fired={summary['approach_fired']} stage2={summary['stage2_ready']}"
    )
    print(f"  labels={summary['by_label']}")
    if unmatched_tags:
        print(f"  tagged but no cache: {unmatched_tags}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
